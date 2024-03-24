import tkinter as tk
from tkinter import filedialog
from tkinter import ttk  
import openpyxl
import pyautogui
from time import sleep

def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, file_path)
    set_combobox_options(file_path)

def set_combobox_options(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        first_row_data = []
        for cell in sheet[1]:
            first_row_data.append(str(cell.value))
        column_combobox['values'] = first_row_data
        workbook.close()
    except Exception as e:
        status_label.config(text=f"Erro ao ler a primeira linha: {str(e)}", fg="red")
        column_combobox['values'] = []

def on_filter_change(*args):
    filter_text = filter_entry.get()
    if filter_text:
        filtered_values = [value for value in column_combobox['values'] if filter_text.lower() in value.lower()]
        column_combobox['values'] = filtered_values
    else:
        set_combobox_options(excel_file_entry.get())

def paste_from_excel():
    excel_file = excel_file_entry.get()
    if not excel_file:
        status_label.config(text="Selecione um arquivo Excel", fg="red")
        return

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        num_rows = sheet.max_row
        selected_column = column_combobox.get()

        if selected_column.isdigit():
            column_index = int(selected_column)
        else:
            column_index = None
            for i, cell in enumerate(sheet[1], start=1):
                if cell.value == selected_column:
                    column_index = i
                    break

        if column_index is None:
            raise ValueError("Coluna não encontrada no arquivo Excel.")

        x_coord = 1382
        y_coord = 177

        x1_coord = 1329
        y1_coord = 363

        pyautogui.click(x_coord, y_coord, duration=0.3)
        sleep(0.5)

        column_count = 0  

        for row in range(2, num_rows + 1):  
            cell_value = sheet.cell(row=row, column=column_index).value
            pyautogui.click(x1_coord, y1_coord, duration=0.3)  
            pyautogui.write(str(cell_value))
            pyautogui.press('enter')  
            sleep(0.7)

            column_count += 1  
            if column_count >= 5:  
                break

        workbook.close()  
        status_label.config(text="Colagem concluída!", fg="green")
    except Exception as e:
        status_label.config(text=f"Erro: {str(e)}", fg="red")

root = tk.Tk()
root.title("AutoTAGs")
root.config(bg='#EDEDED')

excel_file_label = tk.Label(root, text="Arquivo Excel:")
excel_file_label.grid(row=0, column=0, padx=5, pady=5)

excel_file_entry = tk.Entry(root, width=70)
excel_file_entry.grid(row=0, column=1, padx=5, pady=5)

select_button = tk.Button(root, text="Selecionar", command=select_excel_file)
select_button.grid(row=0, column=2, padx=5, pady=5)

filter_label = tk.Label(root, text="Filtro:")
filter_label.grid(row=1, column=0, padx=5, pady=5)

filter_entry = tk.Entry(root, width=30)
filter_entry.grid(row=1, column=1, padx=5, pady=5)
filter_entry.bind("<KeyRelease>", on_filter_change)

column_label = tk.Label(root, text="Coluna:")
column_label.grid(row=2, column=0, padx=5, pady=5)

column_combobox = ttk.Combobox(root, width=80)
column_combobox.grid(row=2, column=1, padx=5, pady=5)

paste_button = tk.Button(root, text="Colar", command=paste_from_excel)
paste_button.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

status_label = tk.Label(root, text="", fg="black")
status_label.grid(row=3, column=2, columnspan=1, padx=5, pady=5)

root.mainloop()
