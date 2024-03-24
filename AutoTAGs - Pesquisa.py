import tkinter as tk
from tkinter import filedialog
from tkinter import ttk  # Importando ttk para usar Combobox
import openpyxl
import pyautogui
from time import sleep

# Coleta de dados excel
def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, file_path)
    set_combobox_options(file_path)

def set_combobox_options(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        first_row_data = []
        for cell in sheet[1]:
            first_row_data.append(str(cell.value))
        column_combobox.config(values=first_row_data)
        workbook.close()
    except Exception as e:
        status_label.config(text=f"Erro ao ler a primeira linha: {str(e)}", fg="red")
        column_combobox.config(values=[])

def paste_from_excel():
    excel_file = excel_file_entry.get()
    if not excel_file:
        status_label.config(text="Selecione um arquivo Excel", fg="red")
        return

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        num_rows = sheet.max_row
        selected_column = column_combobox.get()  # Obtendo a coluna selecionada da combobox

        if selected_column.isdigit():
            column_index = int(selected_column)
        else:
            column_index = None
            for i, cell in enumerate(sheet[1], start=1):
                if cell.value == selected_column:
                    column_index = i
                    break

        if column_index is None:
            raise ValueError("Coluna não encontrada no arquivo Excel.")

        # Coordenadas fixas
        x_coord = 1382
        y_coord = 177

        x1_coord = 1329
        y1_coord = 363

        pyautogui.click(x_coord, y_coord, duration=0.3)
        sleep(0.5)

        column_count = 0  # Contador para controlar o número de colunas processadas

        for row in range(2, num_rows + 1):  # Coleta linhas e aplica no local
            cell_value = sheet.cell(row=row, column=column_index).value
            pyautogui.click(x1_coord, y1_coord, duration=0.3)  # clique adicional após a escrita
            pyautogui.write(str(cell_value))
            pyautogui.press('enter')  # Simular pressionar Enter
            sleep(0.7)

            column_count += 1  # Incrementa o contador de colunas processadas
            if column_count >= 5:  # Interrompe o loop após processar 8 colunas
                break

        workbook.close()  # Retorno do loop
        status_label.config(text="Colagem concluída!", fg="green")
    except Exception as e:
        status_label.config(text=f"Erro: {str(e)}", fg="red")

# Janela principal
root = tk.Tk()
root.title("AutoTAGs")
root.config(bg='#EDEDED')  # Cor de fundo geral

# Layout da interface gráfica
excel_file_label = tk.Label(root, text="Arquivo Excel:")
excel_file_label.grid(row=0, column=0, padx=5, pady=5)

excel_file_entry = tk.Entry(root, width=70)
excel_file_entry.grid(row=0, column=1, padx=5, pady=5)

select_button = tk.Button(root, text="Selecionar", command=select_excel_file)
select_button.grid(row=0, column=2, padx=5, pady=5)

column_label = tk.Label(root, text="Coluna:")
column_label.grid(row=1, column=0, padx=5, pady=5)

# Alteração aqui: Usando Combobox para selecionar a coluna
column_combobox = ttk.Combobox(root, width=80)
column_combobox.grid(row=1, column=1, padx=5, pady=5)

paste_button = tk.Button(root, text="Colar", command=paste_from_excel)
paste_button.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

status_label = tk.Label(root, text="", fg="black")
status_label.grid(row=2, column=2, columnspan=1, padx=5, pady=5)

# Executa a interface gráfica
root.mainloop()
