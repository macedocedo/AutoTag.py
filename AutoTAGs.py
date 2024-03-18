import tkinter as tk
from tkinter import filedialog
import openpyxl
import pyautogui
from time import sleep

# Coleta de dados excel
def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, file_path)

def paste_from_excel():
    excel_file = excel_file_entry.get()
    if not excel_file:
        status_label.config(text="Selecione um arquivo Excel primeiro!", fg="red")
        return

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        num_rows = sheet.max_row
        selected_column = column_entry.get()

        if selected_column.isalpha():
            column_index = openpyxl.utils.column_index_from_string(selected_column)
        else:
            column_index = int(selected_column)

        # Coordenadas fixas
        x_coord = 1053
        y_coord = 183

        for row in range(1, num_rows + 1):
            cell_value = sheet.cell(row=row, column=column_index).value

        pyautogui.click(x_coord, y_coord, duration=0.5)
        pyautogui.write(str(cell_value))
        sleep(0.5)
        pyautogui.click(x=1044, y=368, duration=0.3)



        workbook.close()
        status_label.config(text="Colagem concluída!", fg="green")
    except Exception as e:
        status_label.config(text=f"Erro: {str(e)}", fg="red")

# Cria a janela principal
root = tk.Tk()
root.title("Paste from Excel")

# Layout da interface gráfica
excel_file_label = tk.Label(root, text="Arquivo Excel:")
excel_file_label.grid(row=0, column=0, padx=5, pady=5)

excel_file_entry = tk.Entry(root, width=50)
excel_file_entry.grid(row=0, column=1, padx=5, pady=5)

select_button = tk.Button(root, text="Selecionar", command=select_excel_file)
select_button.grid(row=0, column=2, padx=5, pady=5)

column_label = tk.Label(root, text="Coluna:")
column_label.grid(row=1, column=0, padx=5, pady=5)

column_entry = tk.Entry(root)
column_entry.grid(row=1, column=1, padx=5, pady=5)

paste_button = tk.Button(root, text="Colar", command=paste_from_excel)
paste_button.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

status_label = tk.Label(root, text="", fg="black")
status_label.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

# Executa a interface gráfica
root.mainloop()
